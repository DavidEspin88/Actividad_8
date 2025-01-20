from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify, Response
import psycopg2
import csv
import json
from fpdf import FPDF
import pandas as pd
import io
from werkzeug.security import generate_password_hash, check_password_hash
import xml.etree.ElementTree as ET
from datetime import date
from openpyxl import Workbook

from db import get_db_connection


app = Flask(__name__)
app.secret_key = '1234567890'  # Necesario para manejar sesiones

# Ruta para login
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM usuarios WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and user[2] == password:  # Sin bcrypt por simplicidad
            session['user'] = user[1]
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos')

    return render_template('login.html')

# Ruta para dashboard
@app.route('/dashboard')
def dashboard():
    if 'user' in session:
        return render_template('dashboard.html')
    else:
        return redirect(url_for('login'))

# Ruta para listar productos
@app.route('/articulos')
def listar_articulos():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM productos')
    articulos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template('articulos.html', articulos=articulos)

# Ruta para agregar productos
@app.route('/articulos/agregar', methods=['GET', 'POST'])
def agregar_articulo():
    if request.method == 'POST':
        nombre = request.form['nombre']
        categoria = request.form['categoria']
        codigo = request.form['codigo']
        fecha_ingreso = request.form['fecha_ingreso']
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO productos (nombre, categoria, codigo, fecha_ingreso) VALUES (%s, %s, %s, %s)",
                    (nombre, categoria, codigo, fecha_ingreso))
        conn.commit()
        cur.close()
        conn.close()
        return redirect(url_for('listar_articulos'))

    return render_template('agregar_articulo.html')

# Ruta para editar productos
@app.route('/articulos/editar/<int:id>', methods=['GET', 'POST'])
def editar_articulo(id):
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == 'POST':
        nombre = request.form['nombre']
        categoria = request.form['categoria']
        codigo = request.form['codigo']
        fecha_ingreso = request.form['fecha_ingreso']
        cur.execute("UPDATE productos SET nombre = %s, categoria = %s, codigo = %s, fecha_ingreso = %s WHERE id = %s",
                    (nombre, categoria, codigo, fecha_ingreso, id))
        conn.commit()
        return redirect(url_for('listar_articulos'))
    
    cur.execute('SELECT * FROM productos WHERE id = %s', (id,))
    articulo = cur.fetchone()
    cur.close()
    conn.close()

    return render_template('editar_articulo.html', articulo=articulo)

# Ruta para eliminar producto
@app.route('/articulos/eliminar/<int:id>')
def eliminar_articulo(id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('DELETE FROM productos WHERE id = %s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for('listar_articulos'))

# Rutas de Exportación
@app.route('/exportar/<formato>')
def exportar_datos(formato):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM productos")
    productos = cur.fetchall()
    column_names = [desc[0] for desc in cur.description]

    if formato == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for producto in productos:
            pdf.cell(200, 10, txt=f"{producto}", ln=True)
        pdf_file = 'productos.pdf'
        pdf.output(pdf_file)
        return send_file(pdf_file, as_attachment=True)

    elif formato == 'xlsx':
        df = pd.DataFrame(productos, columns=column_names)
        excel_file = 'productos.xlsx'
        df.to_excel(excel_file, index=False)
        return send_file(excel_file, as_attachment=True)

    elif formato == 'csv':
        df = pd.DataFrame(productos, columns=column_names)
        csv_file = 'productos.csv'
        df.to_csv(csv_file, index=False)
        return send_file(csv_file, as_attachment=True)

    elif formato == 'xml':
        root = ET.Element("Productos")
        for producto in productos:
            producto_element = ET.SubElement(root, "Producto")
            for i, col in enumerate(column_names):
                ET.SubElement(producto_element, col).text = str(producto[i])
        tree = ET.ElementTree(root)
        xml_file = 'productos.xml'
        tree.write(xml_file)
        return send_file(xml_file, as_attachment=True)

    elif formato == 'json':
        data = []
        for est in productos:
            producto_dict = {
                "ID": est[0],
                "Nombre": est[1],
                "Categoria": est[2],
                "Código": est[3],
                "Fecha de Ingreso": est[4].strftime('%Y-%m-%d')  # Convertir la fecha a cadena
            }
            data.append(producto_dict)

        json_output = json.dumps(data, indent=4)
        return send_file(io.BytesIO(json_output.encode()), as_attachment=True, download_name='productos.json', mimetype='application/json')

    return redirect(url_for('productos'))

# Ruta para generar reportes
@app.route('/reportes')
def reportes():
    return render_template('reportes.html')


# Cerrar sesión
@app.route('/logout')
def logout():
    session.pop('loggedin', None)
    session.pop('id', None)
    session.pop('username', None)
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/buscar_productos')
def buscar_productos():
    term = request.args.get('term')
    # Aquí debes tener tu conexión a la base de datos
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    # Convertir a un formato JSON adecuado
    result = []
    for producto in productos:
        result.append({
            'id': producto[0],
            'nombre': producto[1],
            'categoria': producto[2],
            'codigo': producto[3],
            'fecha_ingreso': producto[4].isoformat()  # Convertir a formato ISO para JSON
        })

    return jsonify(result)

@app.route('/export/csv')
def exportar_csv():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    # Crear CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Nombre', 'Categoria', 'Código', 'Fecha de ingreso'])
    for producto in productos:
        writer.writerow(producto)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), as_attachment=True, download_name='productos.csv', mimetype='text/csv')


@app.route('/export/xlsx')
def exportar_xlsx():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Productos"

    # Escribir encabezados
    sheet.append(['ID', 'Nombre', 'Categoria', 'Código', 'Fecha de Ingreso'])

    for producto in productos:
        sheet.append(producto)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='productos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/export/pdf')
def exportar_pdf():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Reporte de Productos", ln=True, align='C')

    for producto in productos:
           pdf.cell(200, 10, txt=f"{producto[1]} {producto[2]} - Código: {producto[3]} - Fecha Ingreso: {producto[4]}", ln=True)
    pdf_file = 'productos.pdf'
    pdf.output(pdf_file)
    return send_file(pdf_file, as_attachment=True)


@app.route('/export/json')
def export_json():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    result = [{'id': e[0], 'nombre': e[1], 'categoria': e[2], 'codigo': e[3], 'fecha_ingreso': e[4].isoformat()} for e in productos]

    response = json.dumps(result)
    return Response(response, mimetype='application/json', headers={"Content-Disposition": "attachment;filename=productos.json"})


@app.route('/export/xml')
def exportar_xml():
    term = request.args.get('term')
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM productos WHERE nombre ILIKE %s OR codigo ILIKE %s", (f'%{term}%', f'%{term}%'))
    productos = cursor.fetchall()

    root = ET.Element("productos")
    for producto in productos:
        est_elem = ET.SubElement(root, "producto")
        ET.SubElement(est_elem, "id").text = str(producto[0])
        ET.SubElement(est_elem, "nombre").text = producto[1]
        ET.SubElement(est_elem, "categoria").text = producto[2]
        ET.SubElement(est_elem, "codigo").text = producto[3]
        ET.SubElement(est_elem, "fecha_ingreso").text = producto[4].isoformat()

    output = io.BytesIO()
    tree = ET.ElementTree(root)
    tree.write(output, encoding='utf-8', xml_declaration=True)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='productos.xml', mimetype='application/xml')

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)
